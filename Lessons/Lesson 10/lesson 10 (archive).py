#1
dct = {1:{11:{111:1111}}, 2:{22:{222:2222}}, 3:{33:333}, 4:4444}
for i in dct:
    print(i, dct[i])
    if type(dct[i]) == dict:
        for j in dct[i]:
            print(j, dct[i][j])
            if type(dct[i][j]) == dict:
                for n in dct[i][j]:
                    print(n, dct[i][j][n])
#2
f = open("Text.txt", "w", encoding="UTF-8")
n = int(input())
for i in range(n):
    f.write(f"{i} - {i ** 2}\n")
f.close()
#3
with open("TestIn.txt") as fi:
    lst = []
    for i in fi:
        lst.append(i.strip())
    print(lst)
    with open("testOut.txt", 'w') as fo:
        for line in lst:
            res = sorted(line.split())
            fo.write(' '.join(res) + ('\n'))
#4
with open('test.csv', 'r', encoding='utf-8') as fi:
    data = fi.read()
    for line in data.splitlines():
        print(', '.join(line.split()))
#5
import csv
columns = ['firs_name', 'second_name', 'rate']
data = [['Ivan', 'Ivanov', '123'], ['Piotr', 'Petrov', '234'], ['Stepan', 'Stepanov', '345']]
with open('test2.csv', 'w', encoding='utf-8', newline='') as file:
    wrtr = csv.writer(file)
    wrtr.writerow(columns)
    for row in data:
        wrtr.writerow(row)
#6
import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active

wb.save('test.xlsx')
for cellobj in ws['A1':'C3']:
    for cell in cellobj:
        print(cell.coordinate, cell.value)
    print('---END---')
for i in range(ws.max_row):
    for j in range(ws.max_column):
        print(i + 1, j + 1, ws.cell(row = i + 1, column = j + 1).value)