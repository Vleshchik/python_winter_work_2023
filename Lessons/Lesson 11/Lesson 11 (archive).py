#1
import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('task10_2.xlsx')
for sh in wb.sheetnames:
    ws = wb[sh]
    print(ws.title, '--------------------')
    for i in range(1, ws.max_row +1):
        for j in range(1, ws.max_column +1):
            print(i+1, j+1, ws.cell(i, j).value)
#2
import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('task10_2.xlsx')
sheets = wb.sheetnames
sheet = wb[sheets[2]]
tes = set()
dict = {}
for i in range(sheet.max_row):
    tes.add(sheet.cell(row = i+1, column = 1).value)
lst = list(tes)
for n in range(len(lst)):
    dict.setdefault(lst[n], 0)
for i in range(sheet.max_row):
    s = sheet.cell(row = i + 1, column = 1).value
    dict[s] = dict[s] + sheet.cell(row=i + 1, column=2).value
sheet.append(['ИТОГО', sum(dict.values())])
sheet = wb[sheets[3]]
sheet.append({1:'111', 2:'222', 3:'333'})
wb.save('task10_2.xlsx')
#3
import time
x = 1000000
for i in range(0, x*10, x):
    t0 = time.time()
    su = 0
    for j in range(i):
        su += j
    t1 = time.time()
    print(i, t1 - t0)
#4
import time
d = {}
for i in range(11):
    if i % 2 == 0:
        def sl2():
            t0 = time.time()
            time.sleep(2)
            t1 = time.time()
            return (t1 - t0)
        d.setdefault(i, sl2())
    else:
        def sl3():
            t0 = time.time()
            time.sleep(3)
            t1 = time.time()
            return (t1 - t0)
        d.setdefault(i, sl3())
for i in range(len(d)):
    print(i, d[i])
print(sum(d.values()))
#5
import datetime
from datetime import date
import locale
locale.setlocale(locale.LC_ALL, "ru")
birthday = date(1995, 1, 11)
print('Название месяца:', birthday.strftime('%B'))
print('Название дня недели:', birthday.strftime('%A'))
print('Год:', birthday.strftime('%Y'))
print('Месяц:', birthday.strftime('%m'))
print('День:', birthday.strftime('%d'))
#6
import calendar
year = int(input('Введите год: '))
d = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0, 6:0}
for month in range(1, 13):
    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        print(year, month, day)
        print(calendar.weekday(year, month, day))
        dd = calendar.weekday(year, month, day)
        d[dd] += 1
print(d)