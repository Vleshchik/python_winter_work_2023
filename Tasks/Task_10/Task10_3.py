import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('task10_3.xlsx')
sheets = wb.sheetnames
sheet = wb[sheets[0]]
tes = set()
dict = {}
for i in range(sheet.max_row):
    tes.add(sheet.cell(row = i+1, column = 1).value)
lst = list(tes)
for n in range(len(lst)):
    dict.setdefault(lst[n], 0)
for i in range(sheet.max_row):
    s = sheet.cell(row = i + 1, column = 1).value
    dict[s] = dict[s] + sheet.cell(row=i + 1, column=2).value
print(dict)
mid = sum(dict.values()) / len(dict)
print(mid)
d = sorted(dict.values())
print(d)
if len(dict) % 2 == 0:
    m_pos_1 = (len(d) // 2)
    m_pos_2 = (len(d) // 2 - 1)
    mediana = (d[m_pos_1] + d[m_pos_2]) / 2
else:
    m_pos = len(d) // 2
    mediana = d[m_pos]
print(mediana)
wb.create_sheet("2")
sheet = wb['2']
sheet.cell(row = 1, column = 1).value = 'Минимальное значение: '
sheet.cell(row = 1, column = 2).value = min(dict.values())
sheet.cell(row = 2, column = 1).value = 'Максимальное значение: '
sheet.cell(row = 2, column = 2).value = max(dict.values())
sheet.cell(row = 3, column = 1).value = 'Среднее арифметическое: '
sheet.cell(row = 3, column = 2).value = mid
sheet.cell(row = 4, column = 1).value = 'Медиана: '
sheet.cell(row = 4, column = 2).value = mediana
wb.save('task10_3.xlsx')
