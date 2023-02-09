import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('task10_1.xlsx')
sheets = wb.sheetnames
sheet = wb[sheets[0]]
tes = set()
dict = {}
for i in range(sheet.max_row):
    tes.add(sheet.cell(row = i+1, column = 1).value)
lst = list(tes)
for n in range(len(lst)):
    dict.setdefault(lst[n], 0)
for i in range(sheet.max_row):
    s = sheet.cell(row = i + 1, column = 1).value
    dict[s] = dict[s] + sheet.cell(row=i + 1, column=2).value
dict.setdefault("ИТОГО", sum(dict.values()))
sheet = wb[sheets[1]]
i = 0
for k, v in dict.items():
    sheet.cell(row=i + 1, column=1).value = k
    sheet.cell(row = i + 1, column = 2).value = v
    i += 1
wb.save('task10_1.xlsx')


