import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('task10_2.xlsx')
sheets = wb.sheetnames
tes = set()
dict = {}
for sheet_name in sheets:
    sheet = wb[sheet_name]
    for i in range(sheet.max_row):
        tes.add(sheet.cell(row = i+1, column = 1).value)
lst = sorted(list(tes))
for n in range(len(lst)):
    dict.setdefault(lst[n], 0)
print(dict)
for sheet_name in sheets:
    sheet = wb[sheet_name]
    for i in range(sheet.max_row):
        s = sheet.cell(row=i + 1, column=1).value
        dict[s] = dict[s] + sheet.cell(row=i + 1, column=2).value
print(dict)
wb.create_sheet("3")
sheet = wb['3']
i = 0
for k, v in dict.items():
    sheet.cell(row=i + 1, column=1).value = k
    sheet.cell(row = i + 1, column = 2).value = v
    i += 1
wb.save('task10_2.xlsx')