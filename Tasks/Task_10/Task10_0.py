import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook('task10_0.xlsx')
print(wb.sheetnames)
sheets = wb.sheetnames
count = 0
lst = []
for sheet_name in sheets:
    sheet = wb[sheet_name]
    count = 0
    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            if sheet.cell(row = i + 1, column = j + 1).value != '':
                count += 1
    tuple = (sheet.title, count)
    lst.append(tuple)
print(lst)
print(sorted(lst))
print(sorted(lst, key = lambda x: -x[1]))